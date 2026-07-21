# -*- coding: utf-8 -*-
"""
集装箱装箱计算 · 网页版 (Streamlit + Plotly)
================================================================
运行方式：
    pip install -r requirements.txt
    streamlit run app.py
浏览器会自动打开一个本地网址(默认 http://localhost:8501)。

功能：
  · 上传数据(txt/xlsx)或直接粘贴，设置重量阈值，选择 标准版/增强版
  · 交互式 3D 装箱结果(可旋转/缩放/悬停/按类别显示隐藏)
  · 导出 Excel 装箱清单 与 JSON 方案
================================================================
"""
import io
import json
import numpy as np
import plotly.graph_objects as go
import pandas as pd
import streamlit as st

import packer_pro as pp
from packer_pro import ContainerPacker, MM, to_mm, to_int, to_float


# --------------------------------------------------------------------------- #
#                               数据解析                                       #
# --------------------------------------------------------------------------- #
def parse_txt(text):
    # 跳过空行与整行注释（# 开头），使带注释的导出文件可以原样再导入
    lines = [ln.strip() for ln in text.splitlines()
             if ln.strip() and ln.split('#')[0].strip()]
    container = tuple(to_mm(x) for x in lines[0].split('#')[0].strip().split()[:3])
    box_count = int(lines[1].split('#')[0].strip())
    boxes = []
    for i in range(box_count):
        parts = lines[2 + i].split('#')[0].strip().split()
        if len(parts) < 6:
            raise ValueError(f"第 {3 + i} 行数据列数不足（需要 长 宽 高 数量 重量 类型），实际: {parts}")
        boxes.append((to_mm(parts[0]), to_mm(parts[1]), to_mm(parts[2]),
                      to_int(parts[3]), to_float(parts[4]), parts[5]))
    return container, boxes


def parse_xlsx(file_like):
    import openpyxl
    wb = openpyxl.load_workbook(file_like, data_only=True)
    sheet = wb.active
    container = tuple(to_mm(cell.value) for cell in list(sheet[1])[:3])
    box_count = int(sheet.cell(row=2, column=1).value)
    boxes = []
    for i in range(3, 3 + box_count):
        row = [cell.value for cell in sheet[i]]
        boxes.append((to_mm(row[0]), to_mm(row[1]), to_mm(row[2]),
                      to_int(row[3]), to_float(row[4]), str(row[5])))
    return container, boxes


# --------------------------------------------------------------------------- #
#                       表格 <-> 内部数据 互转                                 #
# --------------------------------------------------------------------------- #
TYPE_LABELS = {'0': '木箱', '1': '纸箱', '2': '托盘'}
TYPE_CODES = {v: k for k, v in TYPE_LABELS.items()}
DF_COLUMNS = ['长(m)', '宽(m)', '高(m)', '数量', '重量(kg)', '类型']


def boxes_to_df(boxes):
    """内部 boxes(mm) -> 供表格展示的 DataFrame(米/中文类型)。"""
    rows = []
    for (l, w, h, n, weight, t) in boxes:
        rows.append({'长(m)': l / MM, '宽(m)': w / MM, '高(m)': h / MM,
                     '数量': int(n), '重量(kg)': float(weight),
                     '类型': TYPE_LABELS.get(str(t), str(t))})
    return pd.DataFrame(rows, columns=DF_COLUMNS)


def df_to_boxes(df):
    """表格 DataFrame -> 内部 boxes(mm)，自动跳过空行/无效行。"""
    boxes = []
    for _, row in df.iterrows():
        l, w, h = to_mm(row['长(m)']), to_mm(row['宽(m)']), to_mm(row['高(m)'])
        n, weight = to_int(row['数量']), to_float(row['重量(kg)'])
        t = TYPE_CODES.get(str(row['类型']).strip(), str(row['类型']).strip())
        if l > 0 and w > 0 and h > 0 and n > 0:
            boxes.append((l, w, h, n, weight, t))
    return boxes


CARGO_TYPES = ['木箱', '纸箱', '托盘']


def _new_row(l=0.0, w=0.0, h=0.0, n=1, wt=0.0, type='木箱'):
    """新建一行货物，分配全局唯一 id（避免删除后控件状态错位）。"""
    rid = st.session_state.get('next_id', 0)
    st.session_state.next_id = rid + 1
    return {'id': rid, 'l': float(l), 'w': float(w), 'h': float(h),
            'n': int(n), 'wt': float(wt), 'type': type}


def sample_rows():
    data = [(0.49, 0.4, 0.09, 44, 18.0, '纸箱'), (1.2, 1.0, 1.35, 7, 1472.0, '木箱'),
            (0.6, 0.26, 0.37, 120, 12.0, '纸箱'), (0.35, 0.28, 0.35, 168, 12.0, '纸箱'),
            (1.1, 1.1, 1.4, 2, 950.0, '木箱')]
    return [_new_row(*d) for d in data]


def boxes_to_rows(boxes):
    return [_new_row(l / MM, w / MM, h / MM, n, weight, TYPE_LABELS.get(str(t), str(t)))
            for (l, w, h, n, weight, t) in boxes]


def rows_to_boxes(rows):
    """货物行 -> 内部 boxes(mm)，自动跳过空行/无效行。"""
    out = []
    for r in rows:
        l, w, h = to_mm(r['l']), to_mm(r['w']), to_mm(r['h'])
        n, wt = to_int(r['n']), to_float(r['wt'])
        t = TYPE_CODES.get(r['type'], r['type'])
        if l > 0 and w > 0 and h > 0 and n > 0:
            out.append((l, w, h, n, wt, t))
    return out


def _valid_rows(rows):
    return [r for r in rows
            if to_mm(r['l']) > 0 and to_mm(r['w']) > 0 and to_mm(r['h']) > 0 and to_int(r['n']) > 0]


def build_cargo_txt(container_mm, rows):
    """把当前录入的集装箱+货物清单导出为 txt。

    格式与导入完全一致（行内 # 之后为注释，导入时会自动忽略），
    因此导出的文件可以直接用「从文件/文本导入」读回来，实现间断录入。
    """
    valid = _valid_rows(rows)
    L, W, H = (c / MM for c in container_mm)
    lines = [f"{L:g} {W:g} {H:g}    # 集装箱 长 宽 高（米）",
             f"{len(valid)}    # 货物种类数",
             "# 以下每行：长 宽 高 数量 重量 类型(0=木箱 1=纸箱 2=托盘)"]
    for r in valid:
        code = TYPE_CODES.get(r['type'], '0')
        label = TYPE_LABELS.get(code, str(r['type']))
        lines.append(f"{float(r['l']):g} {float(r['w']):g} {float(r['h']):g} "
                     f"{int(r['n'])} {float(r['wt']):g} {code}    # {label}")
    return ("\n".join(lines) + "\n").encode('utf-8')


def build_cargo_xlsx(container_mm, rows):
    """把当前录入的集装箱+货物清单导出为 xlsx（同样可直接再导入）。

    第7列的中文类型仅供阅读，导入时会被忽略。
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    valid = _valid_rows(rows)
    L, W, H = (c / MM for c in container_mm)
    wb = Workbook()
    ws = wb.active
    ws.title = '货物清单'
    ws.append([L, W, H, None, None, None, '集装箱 长 宽 高（米）'])
    ws.append([len(valid), None, None, None, None, None, '货物种类数'])
    for r in valid:
        code = TYPE_CODES.get(r['type'], '0')
        ws.append([float(r['l']), float(r['w']), float(r['h']),
                   int(r['n']), float(r['wt']), code, TYPE_LABELS.get(code, '')])
    for idx, width in enumerate([12, 12, 12, 10, 12, 10, 26], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
#                               装箱计算                                       #
# --------------------------------------------------------------------------- #
def run_packing(container, boxes, threshold, mode, budget):
    packer = ContainerPacker(container, boxes, threshold)
    if mode == "enhanced":
        fits, msg, stats = packer.pack_enhanced(time_budget=budget, verbose=False)
    else:
        fits, msg = packer.pack()
        used = sum(l * w * h for _, _, _, l, w, h in packer.packing_plan)
        stats = {
            'placed': len(packer.packing_plan), 'total': packer.total_units,
            'utilization': used / packer.container_volume if packer.container_volume else 0.0,
            'evaluations': 1, 'seconds': 0.0, 'optimal_full': fits,
        }
    return packer, fits, msg, stats


# --------------------------------------------------------------------------- #
#                           交互式 3D (Plotly)                                 #
# --------------------------------------------------------------------------- #
_TAB10 = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
          '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']

# 立方体的 12 个三角面（顶点索引），顶点顺序见 _cuboid_vertices
_TRI_I = [0, 0, 0, 0, 4, 4, 0, 0, 1, 1, 2, 2]
_TRI_J = [1, 2, 1, 4, 5, 6, 3, 4, 5, 2, 6, 3]
_TRI_K = [2, 3, 5, 5, 6, 7, 7, 7, 6, 6, 7, 7]


def _cuboid_vertices(x, y, z, dx, dy, dz):
    return [(x, y, z), (x + dx, y, z), (x + dx, y + dy, z), (x, y + dy, z),
            (x, y, z + dz), (x + dx, y, z + dz), (x + dx, y + dy, z + dz), (x, y + dy, z + dz)]


def build_figure(packer, highlight_categories=None, region=None):
    """构建交互式 3D 图：每类一个网格，可在图例中点选显示/隐藏。

    region 非空时为 (x0, x1, y0, y1, z0, z1)（单位 mm），只绘制"起点"落在该范围内的
    箱子，用于沿装货方向（长/宽）或高度分段查看装载过程。
    """
    if highlight_categories is None:
        highlight_categories = {info['input_order'] + 1 for info in packer.box_colors}

    fig = go.Figure()

    # 集装箱线框
    L, W, H = [c / MM for c in packer.container]
    edges = [((0, 0, 0), (L, 0, 0)), ((0, W, 0), (L, W, 0)), ((0, 0, H), (L, 0, H)), ((0, W, H), (L, W, H)),
             ((0, 0, 0), (0, W, 0)), ((L, 0, 0), (L, W, 0)), ((0, 0, H), (0, W, H)), ((L, 0, H), (L, W, H)),
             ((0, 0, 0), (0, 0, H)), ((L, 0, 0), (L, 0, H)), ((0, W, 0), (0, W, H)), ((L, W, 0), (L, W, H))]
    ex, ey, ez = [], [], []
    for a, b in edges:
        ex += [a[0], b[0], None]
        ey += [a[1], b[1], None]
        ez += [a[2], b[2], None]
    fig.add_trace(go.Scatter3d(x=ex, y=ey, z=ez, mode='lines',
                               line=dict(color='black', width=3), name='集装箱', hoverinfo='skip'))

    # 按类别分组
    cats = {}
    for (x, y, z, l, w, h), info in zip(packer.packing_plan, packer.box_colors):
        cats.setdefault(info['input_order'], []).append(((x, y, z, l, w, h), info))

    for io in sorted(cats):
        category = io + 1
        selected = category in highlight_categories
        color = _TAB10[io % len(_TAB10)] if selected else '#d9d9d9'
        opacity = 0.85 if selected else 0.15
        vx, vy, vz, ii, jj, kk, htext = [], [], [], [], [], [], []
        base = 0
        for (x, y, z, l, w, h), info in cats[io]:
            if region is not None:
                x0, x1, y0, y1, z0, z1 = region
                if not (x0 <= x <= x1 and y0 <= y <= y1 and z0 <= z <= z1):
                    continue  # 分段查看：起点不在所选范围内的箱子暂不显示
            for (vX, vY, vZ) in _cuboid_vertices(x / MM, y / MM, z / MM, l / MM, w / MM, h / MM):
                vx.append(vX); vy.append(vY); vz.append(vZ)
            ii += [base + t for t in _TRI_I]
            jj += [base + t for t in _TRI_J]
            kk += [base + t for t in _TRI_K]
            dl, dw, dh = info['original_dimensions']
            tlabel = TYPE_LABELS.get(str(info['type']), str(info['type']))
            htext.append(f"箱号 {info['number']}｜类别 {category}｜{tlabel}<br>"
                         f"尺寸 {dl:g}×{dw:g}×{dh:g} m｜重量 {info['weight']:g} kg")
            base += 8
        if not vx:
            continue  # 该类别在当前层高下无可显示的箱子
        qty = packer.boxes[io][3]
        dl, dw, dh = cats[io][0][1]['original_dimensions']
        tlabel = TYPE_LABELS.get(str(cats[io][0][1]['type']), str(cats[io][0][1]['type']))
        fig.add_trace(go.Mesh3d(
            x=vx, y=vy, z=vz, i=ii, j=jj, k=kk,
            color=color, opacity=opacity, flatshading=True,
            name=f"类别{category}: {dl:g}×{dw:g}×{dh:g}m ({tlabel}, {qty}件)",
            showlegend=True, hoverinfo='text',
            hovertext=[t for t in htext for _ in range(8)],
        ))

    fig.update_layout(
        scene=dict(
            xaxis_title='长 Length (m)', yaxis_title='宽 Width (m)', zaxis_title='高 Height (m)',
            aspectmode='data',
        ),
        # 图例移到图表下方：顶部让给 Plotly 工具条（缩放/复位/相机/全屏），避免重叠
        margin=dict(l=0, r=0, t=8, b=0), height=700,
        legend=dict(orientation='h', yanchor='top', y=-0.02, xanchor='left', x=0,
                    bgcolor='rgba(255,255,255,0.75)', bordercolor='#E7ECF4', borderwidth=1,
                    font=dict(size=12)),
    )
    return fig


# --------------------------------------------------------------------------- #
#                           导出：Excel / JSON                                 #
# --------------------------------------------------------------------------- #
def build_dataframe(packer):
    """明细装箱清单：按装载顺序列出每一件货的原始规格、摆放朝向与落位坐标。"""
    rows = []
    for (x, y, z, l, w, h), info in sorted(zip(packer.packing_plan, packer.box_colors),
                                           key=lambda it: it[1]['number']):
        dl, dw, dh = info['original_dimensions']
        rows.append({
            '装载顺序': info['number'],
            '类别': info['input_order'] + 1,
            '类型': TYPE_LABELS.get(str(info['type']), str(info['type'])),
            '原始规格(长×宽×高 m)': f"{dl:g}×{dw:g}×{dh:g}",
            '摆放长(m)': l / MM, '摆放宽(m)': w / MM, '摆放高(m)': h / MM,
            '位置X起点(m)': x / MM, '位置Y起点(m)': y / MM, '位置Z起点(m)': z / MM,
            '重量(kg)': info['weight'],
        })
    return pd.DataFrame(rows)


def build_category_summary(packer):
    """分类汇总：每种货物的规格、应装/已装/未装数量与体积重量小计。"""
    placed = {}
    for info in packer.box_colors:
        placed[info['input_order']] = placed.get(info['input_order'], 0) + 1
    rows = []
    for (l, w, h, n, weight, t, io) in packer.boxes:
        p = placed.get(io, 0)
        rows.append({
            '类别': io + 1,
            '类型': TYPE_LABELS.get(str(t), str(t)),
            '规格(长×宽×高 m)': f"{l / MM:g}×{w / MM:g}×{h / MM:g}",
            '应装数量': n, '已装数量': p, '未装数量': n - p,
            '单件重量(kg)': weight,
            '已装重量(kg)': round(weight * p, 1),
            '单件体积(m³)': round(l * w * h / 1e9, 4),
            '已装体积(m³)': round(l * w * h * p / 1e9, 3),
        })
    return pd.DataFrame(rows)


def build_excel_bytes(packer, stats):
    """导出三页 Excel：汇总 / 分类汇总 / 明细装箱清单。"""
    from openpyxl.utils import get_column_letter
    detail = build_dataframe(packer)
    cats = build_category_summary(packer)
    used_vol = sum(l * w * h for _, _, _, l, w, h in packer.packing_plan) / 1e9
    total_weight = sum(info['weight'] for info in packer.box_colors)
    summary = pd.DataFrame({
        '项目': ['集装箱(长×宽×高 m)', '集装箱容积(m³)', '装入件数', '货物总件数', '未装件数',
                 '空间利用率', '已装货物体积(m³)', '已装货物总重(kg)', '搜索评估次数', '搜索用时(s)'],
        '数值': [
            f"{packer.container[0] / MM:g} × {packer.container[1] / MM:g} × {packer.container[2] / MM:g}",
            round(packer.container_volume / 1e9, 3),
            stats['placed'], stats['total'], stats['total'] - stats['placed'],
            f"{stats['utilization'] * 100:.1f}%",
            round(used_vol, 3), round(total_weight, 1),
            stats.get('evaluations', ''), stats.get('seconds', ''),
        ],
    })
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # 「装箱清单」放首位：打开文件即是与页面一致的明细清单
        for name, frame in (('装箱清单', detail), ('分类汇总', cats), ('汇总', summary)):
            frame.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            for idx, col in enumerate(frame.columns, start=1):
                sample = [str(col)] + [str(v) for v in frame[col].head(300)]
                # 中文字符较宽，按 1.7 倍估算列宽
                width = min(max(max(len(s) for s in sample) * 1.7 + 2, 10), 32)
                ws.column_dimensions[get_column_letter(idx)].width = width
            ws.freeze_panes = 'A2'
        writer.book.active = 0  # 打开 Excel 时定位到「装箱清单」
    return buf.getvalue()


def build_json_bytes(packer):
    data = {
        'container': [c / MM for c in packer.container],
        'boxes': [[l / MM, w / MM, h / MM, n, weight, t] for (l, w, h, n, weight, t, i) in packer.boxes],
        'packing_plan': [
            {'position': [p / MM for p in pos[:3]], 'dimensions': [p / MM for p in pos[3:]], 'box_info': info}
            for pos, info in zip(packer.packing_plan, packer.box_colors)],
    }
    return json.dumps(data, indent=4, ensure_ascii=False).encode('utf-8')


def load_plan_into_packer(plan):
    """从已保存的方案 JSON 重建 packer（仅用于查看，不重新计算）。"""
    container = tuple(to_mm(x) for x in plan['container'])
    boxes = [(to_mm(l), to_mm(w), to_mm(h), int(n), float(weight), t)
             for l, w, h, n, weight, t in plan['boxes']]
    packer = ContainerPacker(container, boxes)
    packer.packing_plan = [
        (to_mm(it['position'][0]), to_mm(it['position'][1]), to_mm(it['position'][2]),
         to_mm(it['dimensions'][0]), to_mm(it['dimensions'][1]), to_mm(it['dimensions'][2]))
        for it in plan['packing_plan']]
    packer.box_colors = [it['box_info'] for it in plan['packing_plan']]
    return packer


# --------------------------------------------------------------------------- #
#                        实时预估 / 预警（即时反馈）                           #
# --------------------------------------------------------------------------- #
def compute_estimate(container_mm, boxes, threshold):
    """根据当前集装箱与货物清单，计算体积/重量预估与潜在问题（不依赖界面）。"""
    cont_vol = container_mm[0] * container_mm[1] * container_mm[2]
    cargo_vol = sum(b[0] * b[1] * b[2] * b[3] for b in boxes)
    if cont_vol > 0:
        cd = sorted(container_mm)
        oversized = [b for b in boxes
                     if not all(sorted((b[0], b[1], b[2]))[i] <= cd[i] for i in range(3))]
    else:
        oversized = []  # 集装箱尚未填写，暂不判断超尺寸
    heavy = [b for b in boxes if b[4] > threshold]
    return {
        'pieces': sum(b[3] for b in boxes),
        'cargo_vol_m3': cargo_vol / 1e9,
        'cont_vol_m3': cont_vol / 1e9,
        'ratio': (cargo_vol / cont_vol) if cont_vol > 0 else 0.0,
        'weight': sum(b[4] * b[3] for b in boxes),
        'over_volume': cont_vol > 0 and cargo_vol > cont_vol,
        'oversized_count': len(oversized),
        'heavy_count': len(heavy),
    }


def render_live_estimate(container_mm, boxes, threshold):
    """把实时预估渲染到界面：四个指标 + 体积进度条 + 预警。"""
    est = compute_estimate(container_mm, boxes, threshold)
    section_title(IC_GAUGE, '实时预估')
    m = st.columns(4)
    m[0].metric('货物件数', f"{est['pieces']}")
    m[1].metric('货物总体积', f"{est['cargo_vol_m3']:.2f} m³")
    m[2].metric('体积占集装箱', f"{est['ratio'] * 100:.1f}%")
    m[3].metric('货物总重', f"{est['weight']:.0f} kg")
    st.progress(min(est['ratio'], 1.0))
    if est['over_volume']:
        st.error(f"⚠ 货物总体积 {est['cargo_vol_m3']:.2f} m³ 已超过集装箱 "
                 f"{est['cont_vol_m3']:.2f} m³，必然装不下。")
    if est['oversized_count']:
        st.warning(f"⚠ 有 {est['oversized_count']} 类货物任意摆放都超过集装箱内尺寸，无法装入。")
    if est['heavy_count']:
        st.info(f"ℹ 有 {est['heavy_count']} 类为重货(>{threshold:g}kg)，只能放底层、不可被叠压。")


# --------------------------------------------------------------------------- #
#                          界面主题（视觉设计）                                #
# --------------------------------------------------------------------------- #
THEME_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"], .stApp, button, input, textarea, select {
    font-family: 'Inter', 'Microsoft YaHei', -apple-system, sans-serif;
}
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header[data-testid="stHeader"] {background: transparent;}
html {background: #EDF2FA;}
.stApp {background: transparent;}
.block-container {position: relative; z-index: 1; padding-top: 1.5rem;
    padding-bottom: 4.5rem; max-width: 1240px;}

/* 动态极光背景 */
.aurora {position: fixed; inset: 0; z-index: 0; overflow: hidden; pointer-events: none;}
.aurora .blob {position: absolute; width: 540px; height: 540px; border-radius: 50%;
    filter: blur(74px); opacity: .26; animation: drift 24s ease-in-out infinite;}
.aurora .b1 {background: #3B82F6; top: -150px; left: -90px;}
.aurora .b2 {background: #22D3EE; top: 4%; right: -150px; animation-delay: -8s;}
.aurora .b3 {background: #8B5CF6; bottom: -170px; left: 24%; animation-delay: -14s;}
.aurora .b4 {background: #38BDF8; bottom: -90px; right: 16%; animation-delay: -4s;}
@keyframes drift {0%,100%{transform: translate(0,0) scale(1);}
    33%{transform: translate(48px,-34px) scale(1.09);}
    66%{transform: translate(-36px,28px) scale(.95);}}

/* Hero 头图 */
.hero {
    background: linear-gradient(120deg, #1E3A8A 0%, #2563EB 50%, #0EA5E9 100%);
    background-size: 185% 185%; animation: heroShift 15s ease infinite;
    border-radius: 22px; padding: 30px 36px; margin-bottom: 22px;
    box-shadow: 0 18px 42px rgba(37,99,235,.30); position: relative; overflow: hidden;
}
@keyframes heroShift {0%{background-position:0% 50%;}50%{background-position:100% 50%;}
    100%{background-position:0% 50%;}}
.hero-title {color:#fff; font-size: 30px; font-weight: 700; margin: 0; letter-spacing:.5px;}
.hero-sub {color:#DBEAFE; font-size: 14.5px; margin: 8px 0 0 0;}
.hero-badges {margin-top: 14px;}
.hero-badge {display:inline-block; background: rgba(255,255,255,.16); color:#EAF2FF;
    border:1px solid rgba(255,255,255,.28); border-radius: 999px;
    padding: 4px 13px; font-size: 12.5px; margin-right: 8px;}

/* 卡片容器 */
div[data-testid="stVerticalBlockBorderWrapper"] {
    background:#fff; border:1px solid #E7ECF4 !important; border-radius: 18px;
    box-shadow: 0 8px 26px rgba(15,23,42,.06); margin-bottom: 18px;
    transition: box-shadow .25s ease, border-color .25s ease;
    animation: fadeUp .5s ease both;
}
div[data-testid="stVerticalBlockBorderWrapper"]:hover {
    box-shadow: 0 14px 38px rgba(15,23,42,.10); border-color: #D7E1F1 !important;}
div[data-testid="stVerticalBlockBorderWrapper"] > div {padding: 10px 20px 14px 20px;}
@keyframes fadeUp {from{opacity:0; transform: translateY(12px);} to{opacity:1; transform:none;}}

/* 指标卡 */
div[data-testid="stMetric"] {
    background: #FFFFFF; border: 1px solid #EDF1F7; border-radius: 14px;
    padding: 14px 16px; box-shadow: 0 3px 12px rgba(15,23,42,.045);
}
div[data-testid="stMetricValue"] {color:#1E3A8A; font-weight: 700;}
div[data-testid="stMetricLabel"] {color:#64748B; font-weight: 600;}

/* 按钮 */
.stButton>button {
    border-radius: 11px; font-weight: 600; border: 1px solid #E2E8F0; padding: 8px 16px;
    transition: all .15s ease;
}
.stButton>button:hover {border-color:#2563EB; color:#2563EB; transform: translateY(-1px);
    box-shadow: 0 4px 14px rgba(37,99,235,.12);}
.stButton>button[kind="primary"] {
    background: linear-gradient(90deg,#2563EB 0%, #1D4ED8 100%); border: none; color:#fff;
    box-shadow: 0 8px 20px rgba(37,99,235,.35);
}
.stButton>button[kind="primary"]:hover {color:#fff; transform: translateY(-1px);
    box-shadow: 0 10px 24px rgba(37,99,235,.45);}

/* 章节标题 */
h3, h4 {color:#0F172A; font-weight: 700;}
.sec-title {display:flex; align-items:center; gap:11px; font-size:18px; font-weight:700;
    color:#0F172A; margin: 2px 0 12px 0;}
.sec-ic {display:inline-flex; width:32px; height:32px; border-radius:10px;
    align-items:center; justify-content:center; color:#fff;
    background: linear-gradient(135deg,#2563EB,#0EA5E9); box-shadow:0 6px 14px rgba(37,99,235,.30);}
.sec-ic svg {width:18px; height:18px;}
.grid-head {font-size:12.5px; font-weight:700; color:#64748B; padding:2px 4px 6px 4px;}

/* 侧边栏 */
section[data-testid="stSidebar"] {background:#FFFFFF; border-right:1px solid #E7ECF4;}

/* 主按钮的矢量图标（闪电=开始计算），替代 emoji */
.stButton>button[kind="primary"]::before {
    content: ""; display: inline-block; width: 17px; height: 17px;
    margin-right: 9px; vertical-align: -3px;
    background: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23ffffff' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M13 2 4 14h7l-1 8 9-12h-7z'/%3E%3C/svg%3E") no-repeat center/contain;
}

/* 计算中的动画：堆叠方块 + 流光文字 */
.calc-loader {display: flex; flex-direction: column; align-items: center; gap: 16px;
    padding: 34px 0 28px 0;}
.cubes {display: flex; gap: 11px; align-items: flex-end; height: 46px;}
.cube {width: 20px; border-radius: 5px;
    background: linear-gradient(160deg, #60A5FA, #2563EB);
    box-shadow: 0 6px 14px rgba(37,99,235,.32);
    animation: stack 1.15s ease-in-out infinite;}
.cube:nth-child(1) {height: 20px; animation-delay: 0s;}
.cube:nth-child(2) {height: 30px; animation-delay: .13s;}
.cube:nth-child(3) {height: 40px; animation-delay: .26s;}
.cube:nth-child(4) {height: 30px; animation-delay: .39s;}
.cube:nth-child(5) {height: 20px; animation-delay: .52s;}
@keyframes stack {
    0%, 100% {transform: translateY(0) scaleY(1); opacity: .55;}
    50% {transform: translateY(-13px) scaleY(1.12); opacity: 1;}
}
.calc-text {font-weight: 600; font-size: 15px; letter-spacing: .5px;
    background: linear-gradient(90deg, #1E3A8A, #38BDF8, #1E3A8A);
    background-size: 220% auto; -webkit-background-clip: text; background-clip: text;
    color: transparent; animation: shine 2.2s linear infinite;}
@keyframes shine {to {background-position: 220% center;}}
.calc-sub {font-size: 12px; color: #8091A7;}

/* 页脚签名（随内容排布，永不被遮挡） */
.site-footer {margin: 36px 0 4px 0;}
.footer-line {height: 1px; margin-bottom: 18px;
    background: linear-gradient(90deg, rgba(37,99,235,0), rgba(37,99,235,.35), rgba(37,99,235,0));}
.footer-inner {display: flex; align-items: center; justify-content: center; gap: 14px;}
.footer-txt {line-height: 1.4;}
.footer-name {font-weight: 700; font-size: 14px; color: #0F172A; letter-spacing: .3px;}
.footer-sub {font-size: 11.5px; color: #8091A7; letter-spacing: .5px;}
.sig-mark {font-weight: 700; font-size: 17px; letter-spacing: 1px; color: #fff;
    width: 46px; height: 46px; border-radius: 14px;
    display: flex; align-items: center; justify-content: center;
    background: linear-gradient(135deg, #1E3A8A 0%, #2563EB 55%, #0EA5E9 100%);
    box-shadow: 0 10px 24px rgba(37,99,235,.34), inset 0 1px 0 rgba(255,255,255,.35);
    border: 1px solid rgba(255,255,255,.30);
    transition: transform .25s ease, box-shadow .25s ease;}
.sig-mark:hover {transform: translateY(-2px) rotate(-3deg);
    box-shadow: 0 14px 30px rgba(37,99,235,.45), inset 0 1px 0 rgba(255,255,255,.4);}
</style>
"""

FOOTER_HTML = """
<div class="site-footer">
  <div class="footer-line"></div>
  <div class="footer-inner">
    <span class="sig-mark">HYT</span>
    <div class="footer-txt">
      <div class="footer-name">Designed &amp; Built by HYT</div>
      <div class="footer-sub">智能集装箱装箱系统 · CONTAINER LOADING OPTIMIZER</div>
    </div>
  </div>
</div>
"""

LOADING_HTML = """
<div class="calc-loader">
  <div class="cubes">
    <div class="cube"></div><div class="cube"></div><div class="cube"></div>
    <div class="cube"></div><div class="cube"></div>
  </div>
  <div class="calc-text">正在计算最优装箱方案…</div>
  <div class="calc-sub">正在尝试多种装载策略与货物朝向，请稍候</div>
</div>
"""

HERO_HTML = """
<div class="hero">
  <p class="hero-title"><svg viewBox="0 0 24 24" fill="none" stroke="#fff" stroke-width="1.8" stroke-linejoin="round" style="width:26px;height:26px;vertical-align:-5px;margin-right:9px;"><path d="M12 2 3 7v10l9 5 9-5V7z"/><path d="M3 7l9 5 9-5"/><path d="M12 12v10"/></svg>智能集装箱装箱系统</p>
  <p class="hero-sub">精确装箱 · 遗传搜索逼近最优 · 交互式 3D 可视化 · 一键导出方案</p>
  <div class="hero-badges">
    <span class="hero-badge">标准版 / 增强版双模式</span>
    <span class="hero-badge">实时体积 · 超重预警</span>
    <span class="hero-badge">分段查看装载过程</span>
    <span class="hero-badge">Excel / JSON 导出</span>
  </div>
</div>
"""


AURORA_HTML = ('<div class="aurora"><div class="blob b1"></div><div class="blob b2"></div>'
               '<div class="blob b3"></div><div class="blob b4"></div></div>')

# 章节图标（内联 SVG，精致线性图标，替代简陋 emoji）
IC_BOX = ('<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" '
          'stroke-linejoin="round"><path d="M12 2 3 7v10l9 5 9-5V7z"/><path d="M3 7l9 5 9-5"/>'
          '<path d="M12 12v10"/></svg>')
IC_LIST = ('<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" '
           'stroke-linecap="round"><path d="M8 6h13M8 12h13M8 18h13"/><circle cx="3.5" cy="6" r="1.1"/>'
           '<circle cx="3.5" cy="12" r="1.1"/><circle cx="3.5" cy="18" r="1.1"/></svg>')
IC_GAUGE = ('<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" '
            'stroke-linecap="round"><path d="M4 14a8 8 0 0 1 16 0"/><path d="M12 14l3.5-2.5"/></svg>')
IC_CHART = ('<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" '
            'stroke-linecap="round" stroke-linejoin="round"><path d="M4 20V4M4 20h16"/>'
            '<rect x="7" y="11" width="3" height="6" rx="1"/><rect x="13" y="7" width="3" height="10" rx="1"/></svg>')
IC_CUBE = IC_BOX


def inject_theme():
    st.markdown(THEME_CSS, unsafe_allow_html=True)
    st.markdown(AURORA_HTML, unsafe_allow_html=True)  # 动态极光背景


def render_footer():
    """页脚签名：随内容排布，不会被 Streamlit 悬浮控件遮挡。"""
    st.markdown(FOOTER_HTML, unsafe_allow_html=True)


def render_hero():
    st.markdown(HERO_HTML, unsafe_allow_html=True)


def section_title(icon_svg, text):
    st.markdown(f'<div class="sec-title"><span class="sec-ic">{icon_svg}</span><span>{text}</span></div>',
                unsafe_allow_html=True)


def render_guide():
    """使用说明：这是什么 + 怎么用。默认收起，不干扰熟练用户。"""
    with st.expander('使用说明 —— 这是什么、怎么用（首次使用请先看这里）',
                     expanded=False, icon=':material/menu_book:'):
        st.markdown("""
#### 这是什么
一个**集装箱三维装箱计算工具**。你告诉它集装箱的尺寸和要装的货物（尺寸、数量、重量、类型），
它会自动算出**每一件货放在哪里、怎么摆**，让空间利用率尽可能高，并给出三维图和可下载的装箱清单。

计算时会自动遵守两条实际装载规则：
- **重量规则**：超过你设定「重量阈值」的重货，只能放在**底层地面**，不会被叠到别的货物上面。
- **支撑规则**：任何不落地的货物，底面至少要有 **60%** 被下方货物支撑，避免悬空。

---

#### 怎么用（5 步）

1. **填集装箱尺寸**：在「集装箱尺寸」填长、宽、高（单位：米）。常见 20 尺柜内径约 `5.9 × 2.35 × 2.39`。
2. **填货物清单**：在表格里逐行填 长/宽/高/数量/重量，类型用下拉选**木箱 / 纸箱 / 托盘**。
   - 每行最右侧 🗑 可删除该行；下方有 **添加一行 / 清空清单 / 载入示例**。
   - 也可展开「**从文件 / 文本导入**」，上传旧的 txt/xlsx 或粘贴文本批量导入。
   - **录一半要走？** 展开「**导出当前清单**」，选 **txt** 或 **xlsx** 导出到本地；
     下次再用「**从文件 / 文本导入**」读回来接着填，**不用重新录入**。
3. **设参数**（左侧栏）：
   - **重量阈值(kg)**：超过它的货物不允许被叠压。比如填 100，则 950kg 的托盘只能落地。
   - **计算模式**：**标准版**快速稳定；**增强版**用遗传搜索反复优化，装载率更高但更慢（可设搜索时间上限）。
4. **点「开始计算」**，等待结果。
5. **看结果 & 导出**：查看装载率、三维图与清单，底部可下载 **Excel 装箱清单** 与 **JSON 方案**。

---

#### 结果怎么看
- **实时预估**：还没点计算时就会显示件数、总体积、占集装箱比例、总重，并预警「体积超箱 / 单件超尺寸 / 重货只能放底层」。
- **三维图**：可用鼠标**旋转、缩放**；悬停可看单件详情。
- **类别按钮**：点「类别1·纸箱」这类按钮可切换高亮（蓝色=显示中），也可用「全选 / 全不选」。
- **分段查看**：长/宽/高三个方向各有一个**两端滑块**，可圈定只显示某一段货物。
  实际装柜是沿**长度方向**从一端装到另一端，所以拖动「长度方向」的左端或右端，
  就能模拟**从前往后**或**从后往前**的装载过程；「重置显示范围」一键还原。
- **导出**：Excel 含三页——装箱清单（每件货的原始规格、摆放朝向、落位坐标）、分类汇总、总览。

---

#### 常见问题
- **货物太多，一次录不完？** 随时展开「导出当前清单」存成 txt 或 xlsx，下次导入即可接着录，进度不丢。
  导出的文件带中文注释但**可以原样再导入**，也能用记事本/Excel 直接编辑。
- **想回看以前算过的方案？** 左侧「查看已保存方案」上传之前下载的 `packing_plan.json`，点「载入方案查看」即可重现，无需重算。
- **提示装不下怎么办？** 程序会尽量多装并告诉你装了多少、利用率多少。可尝试：改用增强版、加大集装箱尺寸、或减少货量。
- **长宽填反了有影响吗？** 没有。程序会自动尝试两种朝向并取更优结果。
""")


# --------------------------------------------------------------------------- #
#                       货物清单（可逐行删除的表格）                           #
# --------------------------------------------------------------------------- #
def _delete_row(rid):
    st.session_state.rows = [r for r in st.session_state.rows if r['id'] != rid]


@st.dialog('删除确认')
def _confirm_delete(rid):
    st.write('确定要删除这一行货物吗？删除后不可恢复。')
    skip = st.checkbox('本次使用不再提示删除确认')
    c1, c2 = st.columns(2)
    if c1.button('确认删除', type='primary', use_container_width=True):
        _delete_row(rid)
        if skip:
            st.session_state.skip_del_confirm = True
        st.rerun()
    if c2.button('取消', use_container_width=True):
        st.rerun()


def render_cargo_editor():
    """自定义货物表格：每行可就地编辑，最右侧 🗑 删除该行（含确认与"不再提示"）。"""
    widths = [1.05, 1.05, 1.05, 0.85, 1.05, 1.0, 0.55]
    head = st.columns(widths)
    for col, label in zip(head, ['长(m)', '宽(m)', '高(m)', '数量', '重量(kg)', '类型', '操作']):
        col.markdown(f"<div class='grid-head'>{label}</div>", unsafe_allow_html=True)
    for r in st.session_state.rows:
        rid = r['id']
        c = st.columns(widths)
        r['l'] = c[0].number_input('l', min_value=0.0, value=float(r['l']), step=0.01, format='%.3f',
                                   key=f"l_{rid}", label_visibility='collapsed')
        r['w'] = c[1].number_input('w', min_value=0.0, value=float(r['w']), step=0.01, format='%.3f',
                                   key=f"w_{rid}", label_visibility='collapsed')
        r['h'] = c[2].number_input('h', min_value=0.0, value=float(r['h']), step=0.01, format='%.3f',
                                   key=f"h_{rid}", label_visibility='collapsed')
        r['n'] = c[3].number_input('n', min_value=1, value=int(r['n']), step=1,
                                   key=f"n_{rid}", label_visibility='collapsed')
        r['wt'] = c[4].number_input('wt', min_value=0.0, value=float(r['wt']), step=1.0, format='%.1f',
                                    key=f"wt_{rid}", label_visibility='collapsed')
        idx = CARGO_TYPES.index(r['type']) if r['type'] in CARGO_TYPES else 0
        r['type'] = c[5].selectbox('t', CARGO_TYPES, index=idx, key=f"t_{rid}", label_visibility='collapsed')
        if c[6].button('🗑', key=f"del_{rid}", help='删除此行', use_container_width=True):
            if st.session_state.get('skip_del_confirm'):
                _delete_row(rid)
                st.rerun()
            else:
                _confirm_delete(rid)
    if not st.session_state.rows:
        st.caption('清单为空，点击下方「添加一行」或「载入示例」开始录入。')


# --------------------------------------------------------------------------- #
#                               Streamlit UI                                   #
# --------------------------------------------------------------------------- #
def main():
    st.set_page_config(page_title='智能装箱系统', page_icon='📦', layout='wide')
    inject_theme()
    render_hero()
    render_guide()

    # ---- 会话状态初始化（默认给一份示例，便于上手）----
    st.session_state.setdefault('next_id', 0)
    if 'rows' not in st.session_state:
        st.session_state.rows = []  # 默认空清单，用户可手动添加或点“载入示例”
    for _k in ('cL', 'cW', 'cH'):
        st.session_state.setdefault(_k, None)  # 默认留空

    with st.container(border=True):
        section_title(IC_BOX, '集装箱尺寸（米）')
        cc1, cc2, cc3 = st.columns(3)
        cL = cc1.number_input('长 Length', min_value=0.0, value=st.session_state.cL, step=0.01, format='%.3f')
        cW = cc2.number_input('宽 Width', min_value=0.0, value=st.session_state.cW, step=0.01, format='%.3f')
        cH = cc3.number_input('高 Height', min_value=0.0, value=st.session_state.cH, step=0.01, format='%.3f')
        st.session_state.cL, st.session_state.cW, st.session_state.cH = cL, cW, cH
    container_mm = (to_mm(cL), to_mm(cW), to_mm(cH))

    with st.container(border=True):
        section_title(IC_LIST, '货物清单')
        st.caption('直接在下表修改数值，类型用下拉选择；每行最右侧 🗑 可删除该行（带确认）。')
        render_cargo_editor()
        ac1, ac2, ac3 = st.columns(3)
        if ac1.button('添加一行', icon=':material/add:', use_container_width=True):
            st.session_state.rows.append(_new_row())
            st.rerun()
        if ac2.button('清空清单', icon=':material/delete_sweep:', use_container_width=True):
            st.session_state.rows = []
            st.rerun()
        if ac3.button('载入示例', icon=':material/restart_alt:', use_container_width=True):
            st.session_state.rows = sample_rows()
            st.rerun()

    with st.expander('从文件 / 文本导入（导入后会填入上面的清单供核对）', icon=':material/upload_file:'):
        up = st.file_uploader('上传 txt 或 xlsx', type=['txt', 'xlsx'])
        paste = st.text_area('或粘贴数据文本（第一行集装箱长宽高，第二行种类数，其后每行：长 宽 高 数量 重量 类型）',
                             height=120)
        if st.button('导入到清单'):
            try:
                c = b = None
                if up is not None:
                    if up.name.endswith('.txt'):
                        c, b = parse_txt(up.getvalue().decode('utf-8'))
                    else:
                        c, b = parse_xlsx(io.BytesIO(up.getvalue()))
                elif paste.strip():
                    c, b = parse_txt(paste)
                else:
                    st.warning('请先上传文件或粘贴文本。')
                if c:
                    st.session_state.cL, st.session_state.cW, st.session_state.cH = c[0] / MM, c[1] / MM, c[2] / MM
                    st.session_state.rows = boxes_to_rows(b)
                    st.success('已导入，请在上方核对后点击“开始计算”。')
                    st.rerun()
            except Exception as e:  # noqa
                st.error(f'导入失败：{e}')

    with st.expander('导出当前清单（保存录入进度，下次导入即可继续）', icon=':material/download:'):
        st.caption('导出内容包含集装箱尺寸与全部货物行，可用上方「从文件 / 文本导入」原样读回来继续编辑。')
        ex1, ex2 = st.columns([1, 1])
        fmt = ex1.radio('导出格式', ['txt', 'xlsx'], horizontal=True, key='cargo_export_fmt')
        n_valid = len(_valid_rows(st.session_state.rows))
        if n_valid == 0:
            ex2.button('清单为空，暂无可导出内容', disabled=True, use_container_width=True)
        elif fmt == 'txt':
            ex2.download_button(
                f'导出货物清单（{n_valid} 类 · txt）', icon=':material/download:',
                data=build_cargo_txt(container_mm, st.session_state.rows),
                file_name='货物清单.txt', mime='text/plain', use_container_width=True)
        else:
            ex2.download_button(
                f'导出货物清单（{n_valid} 类 · xlsx）', icon=':material/download:',
                data=build_cargo_xlsx(container_mm, st.session_state.rows),
                file_name='货物清单.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True)

    with st.sidebar:
        section_title(IC_GAUGE, '计算参数')
        threshold = st.number_input('重量阈值 (kg)：超过则不可叠放在其他箱子上', value=100.0, step=10.0)
        mode_label = st.radio('计算模式', ['标准版（快速稳定）', '增强版（搜索逼近最优）'])
        mode = 'enhanced' if mode_label.startswith('增强') else 'standard'
        budget = 60.0
        if mode == 'enhanced':
            budget = st.slider('增强版搜索时间上限（秒）', 10, 600, 60, step=10)
        go_btn = st.button('开始计算', use_container_width=True, type='primary')

        st.divider()
        st.markdown('**查看已保存方案**')
        plan_file = st.file_uploader('导入方案 JSON', type=['json'], label_visibility='collapsed')
        if st.button('载入方案查看', icon=':material/folder_open:', use_container_width=True):
            if plan_file is not None:
                try:
                    plan = json.loads(plan_file.getvalue().decode('utf-8'))
                    pk = load_plan_into_packer(plan)
                    used = sum(l * w * h for _, _, _, l, w, h in pk.packing_plan)
                    st.session_state.result = {
                        'packer': pk, 'fits': len(pk.packing_plan) == pk.total_units, 'msg': '',
                        'mode': 'imported',
                        'stats': {'placed': len(pk.packing_plan), 'total': pk.total_units,
                                  'utilization': used / pk.container_volume if pk.container_volume else 0.0,
                                  'evaluations': '', 'seconds': '', 'optimal_full': True}}
                    st.success('方案已载入，请在右侧查看。')
                    st.rerun()
                except Exception as e:  # noqa
                    st.error(f'方案文件无法解析：{e}')
            else:
                st.warning('请先选择 JSON 方案文件。')

    # ---- 实时预估与预警（随表格即时更新，无需点击计算）----
    boxes = rows_to_boxes(st.session_state.rows)
    with st.container(border=True):
        render_live_estimate(container_mm, boxes, threshold)

    # ---- 触发计算：结果存入会话状态，便于之后调滑块/筛选时保留结果 ----
    if go_btn:
        if min(container_mm) <= 0:
            st.warning('请填写有效的集装箱尺寸（长/宽/高均需大于 0）。')
        elif not boxes:
            st.warning('货物清单为空或没有有效行，请至少填写一行有效货物。')
        else:
            loader = st.empty()
            loader.markdown(LOADING_HTML, unsafe_allow_html=True)  # 计算期间的动画
            try:
                packer, fits, msg, stats = run_packing(container_mm, boxes, threshold, mode, budget)
            finally:
                loader.empty()
            st.session_state.result = {'packer': packer, 'fits': fits, 'msg': msg,
                                       'stats': stats, 'mode': mode}

    res = st.session_state.get('result')
    if not res:
        st.info('填好集装箱尺寸与货物清单后，点击左侧「开始计算」。首次使用可展开顶部「使用说明」。',
                icon=':material/lightbulb:')
        render_footer()
        return

    packer, fits, msg, stats, mode = res['packer'], res['fits'], res['msg'], res['stats'], res['mode']

    st.divider()
    with st.container(border=True):
        section_title(IC_CHART, '装箱结果概览')
        c1, c2, c3, c4 = st.columns(4)
        c1.metric('装入 / 总数', f"{stats['placed']} / {stats['total']}")
        c2.metric('空间利用率', f"{stats['utilization']*100:.1f}%")
        c3.metric('是否全部装入', '是' if fits else '否')
        if mode == 'enhanced':
            c4.metric('搜索评估 / 用时', f"{stats['evaluations']} 次 / {stats['seconds']}s")
        if not fits:
            st.warning(f'未能全部装入：{msg}')

    with st.container(border=True):
        section_title(IC_CUBE, '交互式 3D 装箱结果')
        all_cats = sorted({info['input_order'] + 1 for info in packer.box_colors})
        cat_label = {}
        for info in packer.box_colors:
            cat = info['input_order'] + 1
            if cat not in cat_label:
                tl = TYPE_LABELS.get(str(info['type']), str(info['type']))
                cat_label[cat] = f"类别{cat}·{tl}"
        st.session_state.setdefault('hl', set(all_cats))
        st.session_state.hl = {c for c in st.session_state.hl if c in all_cats}
        st.caption('点击下方按钮切换高亮（蓝色=显示中），也可直接点图例显示/隐藏。')
        tcols = st.columns(len(all_cats) + 2)
        for i, cat in enumerate(all_cats):
            active = cat in st.session_state.hl
            if tcols[i].button(cat_label[cat], key=f'hl_{cat}',
                               type='primary' if active else 'secondary', use_container_width=True):
                if active:
                    st.session_state.hl.discard(cat)
                else:
                    st.session_state.hl.add(cat)
                st.rerun()
        if tcols[len(all_cats)].button('全选', use_container_width=True):
            st.session_state.hl = set(all_cats)
            st.rerun()
        if tcols[len(all_cats) + 1].button('全不选', use_container_width=True):
            st.session_state.hl = set()
            st.rerun()
        L_m = packer.container[0] / MM
        W_m = packer.container[1] / MM
        H_m = packer.container[2] / MM
        st.caption('分段查看：每个方向都可拖动**两端滑块**圈定要显示的区段。实际装柜通常沿'
                   '**长度方向**从一端装到另一端，拖动「长度方向」的左端或右端，即可模拟'
                   '从前往后 / 从后往前的装载过程。')
        r1a, r1b = st.columns(2)
        xr = r1a.slider(f'长度方向 Length（装货方向）  0 ~ {L_m:g} m',
                        0.0, float(L_m), (0.0, float(L_m)),
                        step=max(0.01, round(L_m / 100, 2)), key='rng_x')
        yr = r1b.slider(f'宽度方向 Width  0 ~ {W_m:g} m',
                        0.0, float(W_m), (0.0, float(W_m)),
                        step=max(0.01, round(W_m / 100, 2)), key='rng_y')
        r2a, r2b = st.columns(2)
        zr = r2a.slider(f'高度方向 Height  0 ~ {H_m:g} m',
                        0.0, float(H_m), (0.0, float(H_m)),
                        step=max(0.01, round(H_m / 100, 2)), key='rng_z')
        r2b.markdown('<div style="height:30px"></div>', unsafe_allow_html=True)  # 与滑块对齐
        if r2b.button('重置显示范围', icon=':material/restart_alt:', use_container_width=True):
            for _k in ('rng_x', 'rng_y', 'rng_z'):
                st.session_state.pop(_k, None)
            st.rerun()
        region = (to_mm(xr[0]), to_mm(xr[1]), to_mm(yr[0]), to_mm(yr[1]), to_mm(zr[0]), to_mm(zr[1]))
        st.plotly_chart(build_figure(packer, set(st.session_state.hl), region=region),
                        use_container_width=True,
                        config={'displaylogo': False})

    with st.container(border=True):
        section_title(IC_LIST, '装箱清单')
        st.dataframe(build_dataframe(packer), use_container_width=True, height=320)
        d1, d2 = st.columns(2)
        d1.download_button('下载 Excel 装箱清单', icon=':material/table_view:',
                           data=build_excel_bytes(packer, stats),
                           file_name='packing_plan.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           use_container_width=True)
        d2.download_button('下载 JSON 方案', icon=':material/data_object:',
                           data=build_json_bytes(packer),
                           file_name='packing_plan.json', mime='application/json',
                           use_container_width=True)

    render_footer()


if __name__ == '__main__':
    main()
